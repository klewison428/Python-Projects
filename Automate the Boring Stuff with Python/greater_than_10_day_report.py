import pymysql
import sys
import os
import ntpath
import time
import smtplib
from email import Encoders
from MroConfig import MroConfig
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from datetime import datetime, timedelta
from email.MIMEMultipart import MIMEMultipart
from email.Utils import COMMASPACE, formatdate
from openpyxl import Workbook

MROCONFIG = MroConfig()

MROCONFIG = MroConfig()
BOT_DIR  = os.path.dirname(os.path.realpath(__file__))
time_stamp = time.strftime("%a, %d %b %Y %H:%M:%S +0000", time.gmtime())

with open(os.path.join(BOT_DIR, 'greater_than_10_days_report.log'), 'a') as f:
    f.write('\n[-- greater_than_10_days_report.py @ %s --]\n' % time_stamp)

wb = Workbook()
ws = wb.active
ws.title = "Greater Than 10 Days Report"


def get_greater_than_10_days_report():
	try:
		con = pymysql.connect(MROCONFIG.CONFIG['DB']['host'], MROCONFIG.CONFIG['DB']['username'], MROCONFIG.CONFIG['DB']['password'], MROCONFIG.CONFIG['DB']['database'])
		#con = pymysql.connect(host='192.168.1.206', user='jerrym', password='usmro1', db='uservices')
		cur = con.cursor()
		print('connection successful')

		cur.execute("SELECT DATEDIFF(NOW(), l.collection_date) LabResult_DaysSinceCollection, l.coc_number LabResult_SpecimenID, CASE WHEN h.coch_id IS NULL THEN 'Lab Result' ELSE h.which_queue END 'Status', r.result_out LabResult_Result, e.employee_id LabResult_EmployeeID, CONCAT( e.first_name, ' ', e.last_name ) LabResult_EmployeeName, c.company_name LabResult_Company, d.division_name LabResult_Division, l.collection_date LabResult_CollectionDate, l.date_first_reported LabResult_DateFirstReported_LabReported, MIN(cl.create_datetime) RxReceivedDate FROM labr l JOIN empl e ON l.empl_id = e.empl_id JOIN rslt r ON l.rslt_id = r.rslt_id JOIN divi d ON l.divi_id = d.divi_id JOIN cmpy c ON d.cmpy_id = c.cmpy_id LEFT JOIN coch h ON l.coch_id = h.coch_id LEFT JOIN clbl cl ON h.coch_id = cl.coch_id AND cl.qlbl_id = 5 WHERE h.date_signed IS NULL AND DATEDIFF(NOW(), l.collection_date) >= 10 AND l.coch_id <> - 1 AND c.cmpy_id <> 319 GROUP BY l.labr_id ORDER BY 1 DESC LIMIT 10;")
		records = cur.fetchall()
		header = ['Days Since Collection', 'LabResult_SpecimenID', 'Status', 'Lab Result', 'Employee ID', 'Employee Name', 'Company', 'Division', 'Collection Date', 'Date First Reported', 'Rx Received Date']
		c = 1
		for i in header:
			#ws.cell(1, c, value=i)
			ws.cell(row=1, column=c).value = i
			c += 1

		r = 2
		for record in records:
			lab_result_days_since_collection = record[0]
			lab_result_specimen_id = record[1]
			status = record[2]
			lab_result = record[3]
			employee_id = record[4]
			employee_name = record[5]
			company = record[6]
			division = record[7]
			collection_date = record[8]
			date_first_reported = record[9]
			prescription_date = record[10]

			ws.cell(row=r, column=1).value=lab_result_days_since_collection
			ws.cell(row=r, column=2).value=lab_result_specimen_id
			ws.cell(row=r, column=3).value=status
			ws.cell(row=r, column=4).value=lab_result
			ws.cell(row=r, column=5).value=employee_id
			ws.cell(row=r, column=6).value=employee_name
			ws.cell(row=r, column=7).value=company
			ws.cell(row=r, column=8).value=division
			ws.cell(row=r, column=9).value=collection_date
			ws.cell(row=r, column=10).value=date_first_reported
			ws.cell(row=r, column=11).value=prescription_date
			r += 1
		
		wb.save('report.xls')
		# Use the below examples to copy the workbook into an archive
		#shutil.copy(os.path.join(PSY_DIR, file_name), LAB_DIR)
        #shutil.copy(os.path.join(PSY_DIR, file_name), XML_DIR)

		if con:
			con.close()

	except pymysql.Error as e:
		print('Got error {!r}, errno is {}'.format(e, e.args[0]))

# [Function to send email with attachment]:
def sendMail(to, subject, text, files=[],server="192.168.1.150"):
    fro = "University Services <admin@uservices.com>"

    msg = MIMEMultipart()
    msg['From'] = fro
    msg['To'] = COMMASPACE.join(to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach( MIMEText(text) )

    filename = 'report.xls'
    attachment = open(filename, 'rb')

    part = MIMEBase('application', 'octet-stream')
    part.set_payload( (attachment).read() )
    Encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= "+filename)
    msg.attach(part)

    smtp = smtplib.SMTP(server)
    smtp.sendmail(fro, to, msg.as_string() )
    smtp.close()

if __name__ == "__main__":
	_file_name = get_greater_than_10_days_report()
	# [Send report email to uServices]:
	#sendMail(["LCarr@uservices.com", "breid@uservices.com"],
	sendMail(["kristyna.lewison@disa.com"], "Greater Than 10 Days Report", "Attached Excel sheet lists the lab results that have been pending for more than 10 days.", [_file_name])